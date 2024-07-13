import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook, Workbook
import os

def get_excel_metadata(file_path):
    workbook = load_workbook(filename=file_path, read_only=True)
    sheet = workbook.active
    columns = [cell.value for cell in sheet[1]]
    row_count = sheet.max_row
    workbook.close()
    return columns, row_count

def split_excel_file(input_file, output_dir, columns, start_row, end_row):
    chunk_size = 1000
    base_filename = os.path.splitext(os.path.basename(input_file))[0]
    output_file_template = os.path.join(output_dir, f"{base_filename}_out_{{}}.xlsx")
    part_number = 1
    
    for chunk_start in range(start_row, end_row, chunk_size):
        chunk_end = min(chunk_start + chunk_size, end_row)
        df_chunk = pd.read_excel(
            input_file,
            usecols=columns,
            skiprows=range(1, chunk_start),
            nrows=chunk_end - chunk_start
        )
        output_file = output_file_template.format(part_number)
        df_chunk.to_excel(output_file, index=False)
        part_number += 1


def open_file_dialog():
    filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    entry_input_file.delete(0, tk.END)
    entry_input_file.insert(0, filename)
    update_metadata(filename)

def save_directory_dialog():
    directory = filedialog.askdirectory()
    entry_output_dir.delete(0, tk.END)
    entry_output_dir.insert(0, directory)

def update_metadata(file_path):
    try:
        columns, row_count = get_excel_metadata(file_path)
        for widget in frame_columns.winfo_children():
            widget.destroy()
        column_vars.clear()
        for col in columns:
            var = tk.BooleanVar()
            chk = tk.Checkbutton(frame_columns, text=col, variable=var)
            chk.pack(anchor='w')
            column_vars[col] = var
        label_row_count.config(text=f"Total Rows: {row_count}")
        entry_end_row.delete(0, tk.END)
        entry_end_row.insert(0, str(row_count))
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while reading the file: {e}")

def execute_split():
    input_file = entry_input_file.get()
    output_dir = entry_output_dir.get()
    columns = [col for col, var in column_vars.items() if var.get()]
    start_row = int(entry_start_row.get())
    end_row = int(entry_end_row.get())
    
    if not input_file or not output_dir:
        messagebox.showerror("Error", "Please select input file and output directory.")
        return
    
    try:
        split_excel_file(input_file, output_dir, columns, start_row, end_row)
        messagebox.showinfo("Success", "File split successfully!")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

# Set up the main window
root = tk.Tk()
root.title("Excel File Splitter")

# Input file
tk.Label(root, text="Input File:").grid(row=0, column=0, padx=10, pady=5)
entry_input_file = tk.Entry(root, width=50)
entry_input_file.grid(row=0, column=1, padx=10, pady=5)
tk.Button(root, text="Browse", command=open_file_dialog).grid(row=0, column=2, padx=10, pady=5)

# Output directory
tk.Label(root, text="Output Directory:").grid(row=1, column=0, padx=10, pady=5)
entry_output_dir = tk.Entry(root, width=50)
entry_output_dir.grid(row=1, column=1, padx=10, pady=5)
tk.Button(root, text="Browse", command=save_directory_dialog).grid(row=1, column=2, padx=10, pady=5)

# Columns
tk.Label(root, text="Select Columns:").grid(row=2, column=0, padx=10, pady=5)
canvas = tk.Canvas(root)
frame_columns = tk.Frame(canvas)
scrollbar = tk.Scrollbar(root, orient="vertical", command=canvas.yview)
canvas.configure(yscrollcommand=scrollbar.set)
scrollbar.grid(row=2, column=2, sticky='ns', padx=10, pady=5)
canvas.grid(row=2, column=1, padx=10, pady=5)
canvas.create_window((0,0), window=frame_columns, anchor='nw')
frame_columns.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
column_vars = {}

# Start row
tk.Label(root, text="Start Row:").grid(row=3, column=0, padx=10, pady=5)
entry_start_row = tk.Entry(root, width=50)
entry_start_row.grid(row=3, column=1, padx=10, pady=5)

# End row
tk.Label(root, text="End Row:").grid(row=4, column=0, padx=10, pady=5)
entry_end_row = tk.Entry(root, width=50)
entry_end_row.grid(row=4, column=1, padx=10, pady=5)

# Row count
label_row_count = tk.Label(root, text="Total Rows: ")
label_row_count.grid(row=5, column=0, columnspan=3, pady=5)

# Execute button
tk.Button(root, text="Split File", command=execute_split).grid(row=6, column=0, columnspan=3, pady=20)

# Run the application
root.mainloop()
